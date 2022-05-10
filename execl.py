from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
import pandas as pd

def del_col():
    wb = load_workbook('sample.xlsx') # 엑셀 이름 // 파일 열기 // ※ 엑셀 버전이 안 맞으면 에러남
    ws = wb.active  # 해당 엑셀 활성화
    
    # 삭제할 범위 2~8, 7 | 10~11, 2 // 남길 범위 : 9
    # 참고 ) 삭제할 때마다 col 정보가 갱신됨. (2, 3, 4) ---2 삭제---> (2=3, 3=4) -> (2, 3)
    
    for i in range(1,8):    # Index 2 ~ 8 범위 col 삭제 # 7번 반복  
        ws.delete_cols(2)    
    for i in range(1,3):    # Index 10 ~ 11 범위 col 삭제 # 2번 반복 
        ws.delete_cols(3)    # ★ 이 부분 중요 // 위와 다르게 2가 아닌 3! // 즉, 중간을 Skip하여 내용 남기기!
        
    wb.save('samsample.xlsx')   # 엑셀 파일 저장
        


def find_data():
    data = pd.read_excel('samsample.xlsx')  # xlsx 파일 부르기 // csv가 아닌 excel로 열면 sep 걱정X !
    df = pd.DataFrame(data) # DataFrame화
    
    #--------------------------------------------------------
    # df_min = df['this'].min()   # this 열에서 최소값  
    # df_mean = df['this'].mean()  # this 열에서 평균값  
    # df_max = df['this'].max()   # this 열에서 최대값
    # print(df_min, '\n', df_mean, '\n', df_max, '\n', '='*10)
    
    #--------------------------------------------------------
    # df_min = df['this'].idxmin(axis=1,skipna=True)   # this 열에서 최소값 인덱스
    # df_max = df['this'].idxmax(axis=1,skipna=True)   # this 열에서 최대값 인덱스
    # print(df_min, '\n', df_max, '\n', '='*10)
    
    #--------------------------------------------------------
    # df_min = df['this'].argmin()   # this 열에서 최소값 인덱스
    # df_max = df['this'].argmax()   # this 열에서 최대값 인덱스
    # print(df_min, '\n', df_max, '\n', '='*10)
    
    #--------------------------------------------------------
    
    df_min = df.iloc[df['this'].argmin()]   # this 열에서 최소값
    df_mean = df['this'].mean()  # this 열에서 평균값 
    df_max = df.iloc[df['this'].argmax()]   # this 열에서 최대값
    df_10 = df[df['this'] > 10]   # this 열에서 10 보다 큰 수
    
    # print(df_min, '\n', '\n', df_max, '\n', '='*10)
    # print(df_min[1], '\n',  df_mean, '\n', df_max[1], '\n', '='*10)
    # print(df_10)

    new_file = df.to_excel('samsample.xlsx', index=False) # 엑셀 파일 저장, index 표기 X
    
    return df_min, df_max, df_mean, df_10   
    
def save_sheet(df_min, df_max, df_mean, df_10):
    
    # 엑셀 시트 준비
    try:
        wb = load_workbook('Summary.xlsx')  # 파일이 있는 경우에는 기존 파일 사용
        wb.active  # 해당 엑셀 활성화
    except:
        wb = Workbook() # 아닐 경우, 새로 만들기
    
    ws = wb.create_sheet('Num1',1) # 원하는 이름, 자리에 시트 만들기  
    
    #--------------------------------------------------------
    # 셀에 내용 넣기
    ws['B1'] = 'Time'
    ws['C1'] = 'Value'
        
    ws['A2'] = 'Min'
    ws['A3'] = 'Mean'
    ws['A4'] = 'Max'
    ws['A5'] = '> 10'
    
    ws['B2'] = df_min[0]
    
    ws['B4'] = df_max[0]
    
    ws['C2'] = df_min[1]
    ws['C3'] = df_mean
    ws['C4'] = df_max[1]
    ws['C5'] = len(df_10)
    
    x1 = 0
    y1 = 0
    for x in range(7, len(df_10)+7):    # 5 번째 row(줄)부터  
        for y in range(2, 4):
            ws.cell(row = x, column = y, value = df_10.iloc[x1, y1])
            y1 += 1
        y1 = 0
        x1 += 1

    #--------------------------------------------------------
    # 차트 그리기
    line_val = Reference(ws, min_row = 1, max_row = 4, min_col = 3, max_col = 3)   # 해당 구역 참조
    line_category = Reference(ws, min_row = 2, max_row = 5, min_col = 1, max_col = 1) # category 참조
    line_chart = LineChart() # Line 차트로 만들기
    line_chart.add_data(line_val, titles_from_data = True)   # titles_from_data = True : 계열 이름 O
    # line_chart.title = 'Summary'   # 차트 제목
    line_chart.y_axis.title = 'Value'    # Y 축 제목
    line_chart.set_categories(line_category)    # category 설정
    ws.add_chart(line_chart, 'E2')   # 차트 삽입 위치
    
    
    #--------------------------------------------------------
    # 내용 저장
    wb.save('Summary.xlsx')

del_col()
df_min, df_max, df_mean, df_10 = find_data()
save_sheet(df_min, df_max, df_mean, df_10)


# PermissionError: [Errno 13] Permission denied: 'samsample.csv' <- 경로 or 파일 열려 있음
# pandas.errors.ParserError: Error tokenizing data. C error: Expected 2 fields in line 11, saw 4 <- sep 맞추기


